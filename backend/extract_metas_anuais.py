"""
Extrai dados de metas anuais (Q1-Q3) das planilhas e gera metas_anuais.csv.
Suporta dois layouts:
  - Vertical: headers na linha 2, dados de linha 3 em diante (coluna 2 = Periodo)
  - Horizontal: headers numa linha com "Periodo" repetido, trimestres lado a lado
"""
import openpyxl
import csv
import os
import re

BASE = (
    r"C:\Users\amanda.paula\FCamara Consultoria e Formação"
    r"\FCamara Files - CONTROLADORIA\30. FP&A NOVO\06. Cockpit - Desenvolvimentos"
    r"\NewDashboard\Dados para apuração de metas\Dados para apuracao de metas anuais"
)
OUT = os.path.join(os.path.dirname(__file__), "metas_anuais.csv")

PERIODO_RE = re.compile(r"Q[1-4]Y\d{2}", re.IGNORECASE)

def cell(ws, r, c):
    v = ws.cell(r, c).value
    return v if v is not None else ""

def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def find_header_row_and_sections(ws):
    """
    Encontra a linha de header e as colunas de início de cada seção de trimestre.
    Retorna (header_row, [(col_start, col_periodo, col_bu, col_avaliado, col_posicao,
                           col_meta_tipo, col_peso_meta, col_ws, col_peso_ws,
                           col_realizado, col_meta, col_trigger, col_ating,
                           col_salario, col_qtde, col_apuracao), ...])
    """
    HDR_KEYS = ["Periodo", "BU", "Avaliado"]

    for r in range(1, 10):
        row_vals = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 1)
                    if ws.cell(r, c).value is not None}
        # Find all columns named "Periodo" in this row
        periodo_cols = [c for c in range(1, ws.max_column + 1)
                        if str(ws.cell(r, c).value or "").strip() == "Periodo"]
        if len(periodo_cols) >= 1:
            sections = []
            for pc in periodo_cols:
                # Read header names starting from pc
                hdrs = {}
                for offset in range(0, 20):
                    h = str(ws.cell(r, pc + offset).value or "").strip()
                    if h and h not in hdrs:
                        hdrs[h] = pc + offset
                # Map expected columns; "Meta" appears twice (type & value), handle with index
                meta_cols = [pc + o for o in range(0, 20)
                             if str(ws.cell(r, pc + o).value or "").strip() == "Meta"]
                sections.append({
                    "header_row": r,
                    "Periodo":    hdrs.get("Periodo", pc),
                    "BU":         hdrs.get("BU", pc + 1),
                    "Avaliado":   hdrs.get("Avaliado", pc + 2),
                    "Posicao":    hdrs.get("Posição", hdrs.get("Posicao", pc + 3)),
                    "MetaTipo":   meta_cols[0] if len(meta_cols) >= 1 else pc + 4,
                    "PesoMeta":   hdrs.get("Peso Meta", pc + 5),
                    "WS":         hdrs.get("WS", pc + 6),
                    "PesoWS":     hdrs.get("Peso WS", pc + 7),
                    "Realizado":  hdrs.get("Realizado", pc + 8),
                    "MetaVal":    meta_cols[1] if len(meta_cols) >= 2 else pc + 9,
                    "Trigger":    hdrs.get("Trigger Min", pc + 10),
                    "Ating":      hdrs.get("Atingimento", pc + 11),
                    "Salario":    hdrs.get("Salário", hdrs.get("Salario", pc + 12)),
                    "Qtde":       hdrs.get("Quantidade", pc + 13),
                    "Apuracao":   hdrs.get("Apuração", hdrs.get("Apuracao", pc + 14)),
                })
            return r, sections
    return None, []


def extract_sheet(ws):
    """Extrai todas as linhas de dados de uma aba, retorna lista de dicts."""
    header_row, sections = find_header_row_and_sections(ws)
    if not sections:
        return []

    rows = []
    for sec in sections:
        hr = sec["header_row"]
        for r in range(hr + 1, ws.max_row + 1):
            periodo = str(cell(ws, r, sec["Periodo"])).strip()
            if not PERIODO_RE.match(periodo):
                continue
            avaliado = str(cell(ws, r, sec["Avaliado"])).strip()
            if not avaliado:
                continue
            rows.append({
                "nome":       avaliado,
                "bu":         str(cell(ws, r, sec["BU"])).strip(),
                "posicao":    str(cell(ws, r, sec["Posicao"])).strip(),
                "periodo":    periodo.upper(),
                "meta_tipo":  str(cell(ws, r, sec["MetaTipo"])).strip(),
                "ws":         str(cell(ws, r, sec["WS"])).strip(),
                "peso_ws":    to_float(cell(ws, r, sec["PesoWS"])),
                "realizado":  to_float(cell(ws, r, sec["Realizado"])),
                "meta":       to_float(cell(ws, r, sec["MetaVal"])),
                "trigger":    to_float(cell(ws, r, sec["Trigger"])),
                "atingimento":to_float(cell(ws, r, sec["Ating"])),
                "salario":    to_float(cell(ws, r, sec["Salario"])),
                "apuracao":   to_float(cell(ws, r, sec["Apuracao"])),
            })
    return rows


def main():
    all_rows = []
    for fname in sorted(os.listdir(BASE)):
        if not fname.endswith(".xlsx"):
            continue
        path = os.path.join(BASE, fname)
        wb = openpyxl.load_workbook(path, data_only=True)
        for sh in wb.sheetnames:
            rows = extract_sheet(wb[sh])
            all_rows.extend(rows)
            print(f"  {fname} / {sh}: {len(rows)} linhas")

    # Normaliza BU: Grupo Mult = Logistics
    for r in all_rows:
        if r["bu"].lower() in ("logistics",):
            r["bu"] = "Grupo Mult"

    # Escreve CSV
    fieldnames = ["nome", "bu", "posicao", "periodo", "meta_tipo", "ws", "peso_ws",
                  "realizado", "meta", "trigger", "atingimento", "salario", "apuracao"]
    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(all_rows)

    print(f"\nGerado: {OUT}  ({len(all_rows)} linhas)")


if __name__ == "__main__":
    main()
