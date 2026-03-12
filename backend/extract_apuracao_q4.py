"""
Script para extrair dados da planilha APURACAO Q4 bd para q4 feito no braço.xlsx
Saídas: premissas_pessoas, premissas_pesos_meta, premissas_pesos_ws,
        premissas_triggers, budget_receita, budget_lb, budget_tcv
"""

import openpyxl
import csv
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = (
    "C:/Users/amanda.paula/FCamara Consultoria e Formação/"
    "FCamara Files - CONTROLADORIA/30. FP&A NOVO/06. Cockpit - Desenvolvimentos/"
    "NewDashboard/Dados para apuração de metas/APURACAO Q4 bd para q4 feito no braço.xlsx"
)
OUT_DIR = "C:/Users/amanda.paula/worker-dashboard/backend/"

print("Abrindo workbook (data_only=True)...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
print(f"Abas disponíveis: {wb.sheetnames}\n")


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def cv(ws, row, col):
    """Valor da célula como string limpa."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else f"{v:.6f}".rstrip('0').rstrip('.')
    return str(v).strip()


def write_csv(filepath, headers, rows):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"  -> Salvo: {os.path.basename(filepath)}  ({len(rows)} linhas de dados)")


def sample(rows, n=5):
    for r in rows[:n]:
        print("    ", r)


# ─────────────────────────────────────────────────────────────
# TAREFA 1 – ABA PREMISSAS
# ─────────────────────────────────────────────────────────────
# Estrutura real observada:
#   col 23(W)=Finance/Grupo(ignorar), col 24(X)=Nome/Avaliado, col 25(Y)=Posição,
#   col 26(Z)=Contrato, col 27(AA)=Sal Q1, col 28(AB)=Sal Q2,
#   col 29(AC)=Sal Q3, col 30(AD)=Sal Q4
#   Header na linha 3
#
# Pesos WS: linha 3=header (O:U = cols 15:21), dados nas linhas 4-5
# Triggers: linha 8=header (P:U = cols 16:21), dados nas linhas 9-12
# Pesos Meta Quarter: linha 15=header (C:G = cols 3:7), dados linhas 16-22
# Pesos Meta Anual  : linha 25=header (C:I = cols 3:9), dados linhas 26-30+

def tarefa1():
    print("=" * 60)
    print("TAREFA 1 – PREMISSAS")
    print("=" * 60)
    ws = wb["PREMISSAS"]

    # ── 1a) Pessoas / Salários ──
    print("\n--- 1a) Pessoas/Salários (header linha 3, cols X:AD = 24:30) ---")
    headers_p = ["Nome", "Posicao", "Contrato", "Sal_Q1", "Sal_Q2", "Sal_Q3", "Sal_Q4"]
    rows_p = []
    for r in range(4, ws.max_row + 1):
        nome = cv(ws, r, 24)
        pos  = cv(ws, r, 25)
        cont = cv(ws, r, 26)
        s1   = cv(ws, r, 27)
        s2   = cv(ws, r, 28)
        s3   = cv(ws, r, 29)
        s4   = cv(ws, r, 30)
        # Para quando não há mais dados (linha completamente vazia nessa região)
        if not any([nome, pos, cont, s1, s2, s3, s4]):
            break
        if nome:  # só linhas com nome preenchido
            rows_p.append([nome, pos, cont, s1, s2, s3, s4])

    print(f"  Extraídas {len(rows_p)} pessoas")
    sample(rows_p)
    write_csv(os.path.join(OUT_DIR, "premissas_pessoas.csv"), headers_p, rows_p)

    # ── 1b) Pesos por Meta – tabela Quarter (linhas 15-22) e Anual (25-30) ──
    print("\n--- 1b) Pesos por Meta ---")
    # Quarter: cols C:G (3:7), header linha 15
    # Anual  : cols C:I (3:9), header linha 25
    # Vamos unir em um CSV único: Periodo, Posicao, TCV, Receita, MB_pct, MC_pct, ENPS, NPS

    headers_pm = ["Periodo", "Posicao", "TCV", "Receita", "MB_pct", "MC_pct", "ENPS", "NPS"]
    rows_pm = []

    # Quarter (sem ENPS/NPS para algumas posições)
    print("  Quarter (linhas 16-22, cols C:G = 3:7):")
    for r in range(16, 24):
        pos  = cv(ws, r, 3)
        tcv  = cv(ws, r, 4)
        rec  = cv(ws, r, 5)
        mb   = cv(ws, r, 6)
        mc   = cv(ws, r, 7)
        if pos:
            rows_pm.append(["Quarter", pos, tcv, rec, mb, mc, "", ""])
            print(f"    [{r}] {[pos, tcv, rec, mb, mc]}")

    # Anual (com ENPS/NPS, cols C:I = 3:9)
    print("  Anual (linhas 26-32, cols C:I = 3:9):")
    for r in range(26, 34):
        pos  = cv(ws, r, 3)
        tcv  = cv(ws, r, 4)
        rec  = cv(ws, r, 5)
        mb   = cv(ws, r, 6)
        mc   = cv(ws, r, 7)
        enps = cv(ws, r, 8)
        nps  = cv(ws, r, 9)
        if pos:
            rows_pm.append(["Anual", pos, tcv, rec, mb, mc, enps, nps])
            print(f"    [{r}] {[pos, tcv, rec, mb, mc, enps, nps]}")

    print(f"  Total: {len(rows_pm)} linhas de pesos")
    write_csv(os.path.join(OUT_DIR, "premissas_pesos_meta.csv"), headers_pm, rows_pm)

    # ── 1c) Pesos por Workstream (linhas 3-5, cols O:U = 15:21) ──
    print("\n--- 1c) Pesos por Workstream ---")
    headers_ws = ["Periodo", "Total", "Apps", "CloudCyber", "Dados", "Hyper", "Demais"]
    rows_ws = []
    print("  (linhas 4-5, cols O:U = 15:21):")
    for r in range(4, 8):
        periodo = cv(ws, r, 15)
        total   = cv(ws, r, 16)
        apps    = cv(ws, r, 17)
        cloud   = cv(ws, r, 18)
        dados   = cv(ws, r, 19)
        hyper   = cv(ws, r, 20)
        demais  = cv(ws, r, 21)
        if periodo:
            rows_ws.append([periodo, total, apps, cloud, dados, hyper, demais])
            print(f"    [{r}] {[periodo, total, apps, cloud, dados, hyper, demais]}")

    print(f"  Extraídas {len(rows_ws)} linhas WS")
    write_csv(os.path.join(OUT_DIR, "premissas_pesos_ws.csv"), headers_ws, rows_ws)

    # ── 1d) Triggers (linha 8=header, linhas 9-12, cols P:U = 16:21) ──
    print("\n--- 1d) Triggers ---")
    headers_t = ["Meta", "Q1", "Q2", "Q3", "Q4", "FY"]
    rows_t = []
    print("  (linhas 9-14, cols P:U = 16:21):")
    for r in range(9, 16):
        meta = cv(ws, r, 16)
        q1   = cv(ws, r, 17)
        q2   = cv(ws, r, 18)
        q3   = cv(ws, r, 19)
        q4   = cv(ws, r, 20)
        fy   = cv(ws, r, 21)
        if meta:
            rows_t.append([meta, q1, q2, q3, q4, fy])
            print(f"    [{r}] {[meta, q1, q2, q3, q4, fy]}")

    print(f"  Extraídas {len(rows_t)} triggers")
    write_csv(os.path.join(OUT_DIR, "premissas_triggers.csv"), headers_t, rows_t)


# ─────────────────────────────────────────────────────────────
# TAREFAS 2/3/4 – BUDGET RECEITA, LB, TCV
# ─────────────────────────────────────────────────────────────
# Estrutura real (Budget Receita e Budget LB):
#   Header na linha 2
#   col 3  = BS (Vertical/Segmento)
#   col 5  = AE (nome do AE responsável, pode variar por quarter)
#   col 7  = AE Q1, col 8 = AE Q2, col 9 = AE Q3, col 10 = AE Q4
#   col 11 = WS
#   col 14 = Cliente (BU)
#   col 44 = 1Q25, col 45 = 2Q25, col 46 = 3Q25, col 47 = 4Q25
#   col 42 = FY 2025 (total anual)
#   col 2  = Tipo (Receita / Margem)
#   Dados a partir da linha 3
#
# Budget TCV:
#   Header linha 2: col3=Vertical/AE, cols 4-15=meses jan-dez 2025, col16=Total
#   col 2 = AE (nome), dados partir da linha 3 (algumas linhas são subtotais)
#   Quarters: Jan+Fev+Mar=Q1, Abr+Mai+Jun=Q2, Jul+Ago+Set=Q3, Out+Nov+Dez=Q4

def fmt_num(v):
    """Formata número como string sem excesso de casas decimais."""
    if v is None or (isinstance(v, str) and (v.strip() == "" or v.strip().startswith("#"))):
        return ""
    try:
        f = float(v)
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return str(v).strip()


def extract_budget_receita_lb(ws_name, out_file, task_label):
    print("=" * 60)
    print(task_label)
    print("=" * 60)

    if ws_name not in wb.sheetnames:
        print(f"  Aba '{ws_name}' não encontrada.")
        return

    ws = wb[ws_name]
    print(f"  Aba: '{ws_name}'")

    # Mapeamento fixo conforme exploração
    # col 3=BS, col 5=AE (geral), col 7=AE_Q1, col 8=AE_Q2, col 9=AE_Q3, col 10=AE_Q4
    # col 11=WS, col 14=Cliente, col 2=Tipo
    # col 44=Q1_25, col 45=Q2_25, col 46=Q3_25, col 47=Q4_25, col 42=FY_25

    headers = ["tipo", "bs", "ae_q1", "ae_q2", "ae_q3", "ae_q4",
               "ws", "cliente", "q1", "q2", "q3", "q4", "fy"]
    rows_out = []

    for r in range(3, ws.max_row + 1):
        tipo    = cv(ws, r, 2)
        bs      = cv(ws, r, 3)
        ae_q1   = cv(ws, r, 7)
        ae_q2   = cv(ws, r, 8)
        ae_q3   = cv(ws, r, 9)
        ae_q4   = cv(ws, r, 10)
        ws_col  = cv(ws, r, 11)
        cliente = cv(ws, r, 14)

        # Valores: cols 44-47 para quarters 2025, col 42 para FY
        raw_q1 = ws.cell(r, 44).value
        raw_q2 = ws.cell(r, 45).value
        raw_q3 = ws.cell(r, 46).value
        raw_q4 = ws.cell(r, 47).value
        raw_fy = ws.cell(r, 42).value

        q1 = fmt_num(raw_q1)
        q2 = fmt_num(raw_q2)
        q3 = fmt_num(raw_q3)
        q4 = fmt_num(raw_q4)
        fy = fmt_num(raw_fy)

        # Ignora linhas completamente vazias
        if not any([tipo, bs, ae_q1, cliente, q1, q2, q3, q4, fy]):
            continue

        rows_out.append([tipo, bs, ae_q1, ae_q2, ae_q3, ae_q4,
                         ws_col, cliente, q1, q2, q3, q4, fy])

    print(f"\n  Colunas: {headers}")
    print(f"  Extraídas {len(rows_out)} linhas")
    print("  Sample (primeiras 5):")
    sample(rows_out)
    write_csv(os.path.join(OUT_DIR, out_file), headers, rows_out)


def extract_budget_tcv():
    print("=" * 60)
    print("TAREFA 4 – BUDGET TCV")
    print("=" * 60)

    ws_name = "Budget TCV"
    if ws_name not in wb.sheetnames:
        print(f"  Aba '{ws_name}' não encontrada.")
        return

    ws = wb[ws_name]
    print(f"  Aba: '{ws_name}'")

    # Estrutura: header linha 2
    # col 2 = AE (nome), col 3 = Vertical/Descrição
    # cols 4-15 = Jan a Dez 2025 (meses mensais)
    # col 16 = Total FY
    # Q1=cols4+5+6, Q2=cols7+8+9, Q3=cols10+11+12, Q4=cols13+14+15

    print("  Verificando linhas 2-8:")
    for r in range(2, 9):
        vals = [(c, cv(ws, r, c)) for c in range(1, 20) if cv(ws, r, c)]
        if vals:
            print(f"    Linha {r}: {vals}")

    headers = ["ae", "descricao", "q1", "q2", "q3", "q4", "fy"]
    rows_out = []

    def sum_cols(ws, r, col_start, col_end):
        total = 0.0
        any_val = False
        for c in range(col_start, col_end + 1):
            v = ws.cell(r, c).value
            if v is not None and not isinstance(v, str):
                try:
                    total += float(v)
                    any_val = True
                except (ValueError, TypeError):
                    pass
        return f"{total:.2f}" if any_val else ""

    for r in range(3, ws.max_row + 1):
        ae    = cv(ws, r, 2)
        desc  = cv(ws, r, 3)

        q1 = sum_cols(ws, r, 4, 6)
        q2 = sum_cols(ws, r, 7, 9)
        q3 = sum_cols(ws, r, 10, 12)
        q4 = sum_cols(ws, r, 13, 15)
        fy = fmt_num(ws.cell(r, 16).value)

        if not any([ae, desc, q1, q2, q3, q4, fy]):
            continue

        rows_out.append([ae, desc, q1, q2, q3, q4, fy])

    print(f"\n  Colunas: {headers}")
    print(f"  Extraídas {len(rows_out)} linhas")
    print("  Sample (primeiras 5):")
    sample(rows_out)
    write_csv(os.path.join(OUT_DIR, "budget_tcv.csv"), headers, rows_out)


# ─────────────────────────────────────────────────────────────
# EXECUÇÃO
# ─────────────────────────────────────────────────────────────

tarefa1()
extract_budget_receita_lb("Budget Receita", "budget_receita.csv", "TAREFA 2 – BUDGET RECEITA")
extract_budget_receita_lb("Budget LB",      "budget_lb.csv",      "TAREFA 3 – BUDGET LB")
extract_budget_tcv()

print("\n" + "=" * 60)
print("CONCLUÍDO")
print("=" * 60)
