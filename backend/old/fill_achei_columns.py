import openpyxl, unicodedata, re, sys
sys.stdout.reconfigure(encoding='utf-8')

def norm(s):
    s = str(s).upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s

path = 'C:/Users/amanda.paula/Downloads/Racional Tentativa yuri.xlsx'
wb = openpyxl.load_workbook(path)

ws_clt  = wb['CUTOS CLTs']
ws_pess = wb['PARA DE PESSOAS']
ws_te   = wb['TimeAndExpenses']

# ── Monta dicionários de lookup (norm -> valor original) ──────────────────

# PARA DE PESSOAS col C
parade_map = {}
for row in ws_pess.iter_rows(min_row=2, min_col=3, max_col=3):
    v = row[0].value
    if v and isinstance(v, str) and v.strip():
        parade_map[norm(v)] = v.strip()

# TimeAndExpenses col L (PROFISSIONAL)
# Descobre qual coluna é L (col 12)
te_map = {}
for row in ws_te.iter_rows(min_row=3, min_col=12, max_col=12):
    v = row[0].value
    if v and isinstance(v, str) and v.strip():
        te_map[norm(v)] = v.strip()

print(f'PARA DE PESSOAS: {len(parade_map)} nomes')
print(f'TimeAndExpenses: {len(te_map)} nomes')

# ── Preenche CUTOS CLTs colunas N e O com valores diretos ─────────────────
col_n = 14  # achei no para de?
col_o = 15  # achei no racional?

n_parade = n_rac = n_miss_parade = n_miss_rac = 0

for row in ws_clt.iter_rows(min_row=2):
    nome_cell = row[3]   # coluna D = índice 3
    v = nome_cell.value
    if not v or not isinstance(v, str) or not v.strip():
        continue

    n = norm(v)
    cell_parade = row[col_n - 1]  # coluna N
    cell_rac    = row[col_o - 1]  # coluna O

    # achei no para de?
    if n in parade_map:
        cell_parade.value = parade_map[n]
        n_parade += 1
    else:
        cell_parade.value = None
        n_miss_parade += 1

    # achei no racional?
    if n in te_map:
        cell_rac.value = te_map[n]
        n_rac += 1
    else:
        cell_rac.value = None
        n_miss_rac += 1

print(f'\nCUTOS CLTs col N (achei no para de?): {n_parade} encontrados, {n_miss_parade} não encontrados')
print(f'CUTOS CLTs col O (achei no racional?): {n_rac} encontrados, {n_miss_rac} não encontrados')

# ── Preenche PARA DE PESSOAS colunas H e I com valores diretos ────────────
col_h = 8  # NOME CLT
col_i = 9  # NOME RACIONAL

# CUTOS CLTs col D: norm -> valor
clt_map = {}
for row in ws_clt.iter_rows(min_row=2, min_col=4, max_col=4):
    v = row[0].value
    if v and isinstance(v, str) and v.strip():
        clt_map[norm(v)] = v.strip()

# CUTOS CLTs col E: ID SAP norm -> valor nome
clt_sap_map = {}
for row in ws_clt.iter_rows(min_row=2, min_col=4, max_col=5):
    nome_v = row[0].value
    sap_v  = row[1].value
    if nome_v and sap_v:
        clt_sap_map[str(sap_v).strip()] = str(nome_v).strip()

h_found = h_miss = i_found = i_miss = 0

for row in ws_pess.iter_rows(min_row=2):
    nome_cell = row[2]   # coluna C = índice 2
    sap_cell  = row[6]   # coluna G (ID SAP num) = índice 6
    v = nome_cell.value
    if not v or not isinstance(v, str) or not v.strip():
        continue

    n = norm(v)
    cell_h = row[col_h - 1]
    cell_i = row[col_i - 1]

    # NOME CLT: busca por nome, fallback por ID SAP
    if n in clt_map:
        cell_h.value = clt_map[n]
        h_found += 1
    elif sap_cell.value and str(sap_cell.value).strip() in clt_sap_map:
        cell_h.value = clt_sap_map[str(sap_cell.value).strip()]
        h_found += 1
    else:
        cell_h.value = None
        h_miss += 1

    # NOME RACIONAL: busca no TimeAndExpenses
    if n in te_map:
        cell_i.value = te_map[n]
        i_found += 1
    else:
        cell_i.value = None
        i_miss += 1

print(f'\nPARA DE PESSOAS col H (NOME CLT): {h_found} encontrados, {h_miss} não encontrados')
print(f'PARA DE PESSOAS col I (NOME RACIONAL): {i_found} encontrados, {i_miss} não encontrados')

wb.save(path)
print('\nSalvo!')
