import openpyxl, unicodedata, re, sys
from rapidfuzz import process, fuzz
sys.stdout.reconfigure(encoding='utf-8')

def norm(s):
    s = str(s).upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s

path = 'C:/Users/amanda.paula/Downloads/Racional Tentativa yuri.xlsx'
wb = openpyxl.load_workbook(path)

ws_clt  = wb['CUTOS CLTs']
ws_pess = wb['PARA DE PESSOAS']

# Coleta todos os nomes de PARA DE PESSOAS col C
# norm -> valor original na célula
pess_map = {}  # norm -> (valor_celula, row)
for row in ws_pess.iter_rows(min_row=2, min_col=3, max_col=3):
    v = row[0].value
    if v and isinstance(v, str) and v.strip():
        n = norm(v)
        if n not in pess_map:
            pess_map[n] = v  # guarda o valor original da célula

print(f'PARA DE PESSOAS: {len(pess_map)} nomes únicos')
pess_norms = list(pess_map.keys())

# Verifica cada nome em CUTOS CLTs col D
# Se não tem match exato → tenta fuzzy → se >=90, ajusta CLT para igualar exatamente o valor de PESS
fixes = 0
sem_match = []

for row in ws_clt.iter_rows(min_row=2, min_col=4, max_col=4):
    cell = row[0]
    v = cell.value
    if not v or not isinstance(v, str) or not v.strip():
        continue
    n = norm(v)

    if n in pess_map:
        # Garante que o valor é idêntico (remove trailing spaces, etc.)
        target = pess_map[n]
        if v != target:
            cell.value = target
            fixes += 1
    else:
        # Tenta fuzzy
        match = process.extractOne(n, pess_norms, scorer=fuzz.token_sort_ratio, score_cutoff=90)
        if match:
            target = pess_map[match[0]]
            print(f'  Fuzzy CLT D{cell.row}: {repr(v)} -> {repr(target)} ({match[1]:.0f})')
            cell.value = target
            fixes += 1
        else:
            sem_match.append((cell.row, v))

print(f'\nAjustes feitos: {fixes}')
print(f'Sem match (genuinamente ausentes de PARA DE PESSOAS): {len(sem_match)}')
for r, v in sem_match:
    print(f'  CLT D{r}: {repr(v)}')

wb.save(path)
print('\nSalvo!')
