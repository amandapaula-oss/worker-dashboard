import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

path = 'C:/Users/amanda.paula/Downloads/Racional Tentativa yuri.xlsx'
wb = openpyxl.load_workbook(path)

ws_clt  = wb['CUTOS CLTs']
ws_pess = wb['PARA DE PESSOAS']

# CUTOS CLTs: restaura fórmulas nas colunas N (14) e O (15)
for row in ws_clt.iter_rows(min_row=2, min_col=4, max_col=4):
    r = row[0].row
    ws_clt.cell(r, 14).value = f"=_xlfn.XLOOKUP(D{r},'PARA DE PESSOAS'!C:C,'PARA DE PESSOAS'!C:C)"
    ws_clt.cell(r, 15).value = f"=_xlfn.XLOOKUP(D{r},TimeAndExpenses!L:L,TimeAndExpenses!L:L)"

print('CUTOS CLTs: fórmulas N e O restauradas')

# PARA DE PESSOAS: restaura fórmulas nas colunas H (8) e I (9)
for row in ws_pess.iter_rows(min_row=2, min_col=3, max_col=3):
    r = row[0].row
    ws_pess.cell(r, 8).value = (
        f"=_xlfn.XLOOKUP(C{r},'CUTOS CLTs'!D:D,'CUTOS CLTs'!D:D,"
        f"_xlfn.XLOOKUP(G{r},'CUTOS CLTs'!E:E,'CUTOS CLTs'!E:E))"
    )
    ws_pess.cell(r, 9).value = f"=_xlfn.XLOOKUP(C{r},TimeAndExpenses!L:L,TimeAndExpenses!L:L)"

print('PARA DE PESSOAS: fórmulas H e I restauradas')

wb.save(path)
print('Salvo!')
